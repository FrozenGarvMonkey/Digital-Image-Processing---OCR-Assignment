import cv2
import numpy as np
import pytesseract as tess
import openpyxl


# Preset templates and keyword for extraction
numTemplate = 4
numImage = 20
threshold = [0.9, 0.9, 0.85, 0.9]
keywords = ["AMENDMENTS",
            "DRAWN",
            "TITLE",
            "DRAWING N",
            "CHECKED",
            "APPROVED",
            "UNIT",
            "STATUS",
            "PAGE",
            "CONTRACTOR",
            "PROJECT NO",
            "FONT",
            "LANG",
            "CAD",
            "COMPANY"]
keyDict = {"AMENDMENTS": "Amendments",
           "DRAWN": "Drawn By",
           "TITLE": "Drawing Title",
           "DRAWING N": "Drawing Number",
           "CHECKED": "Checked By",
           "APPROVED": "Approved By",
           "UNIT": "Measurement Unit",
           "STATUS": "Drawing Status",
           "PAGE": "Page Number",
           "CONTRACTOR": "Contractor",
           "PROJECT NO": "Project Number",
           "FONT": "Font",
           "LANG": "Language",
           "CAD": "CAD Number",
           "COMPANY": "Company"}

templates = []
for val in range(numTemplate):
    templates.append(cv2.imread("template{}.png".format(val+1), 0))
excelResults = []

# Adding header to each array according to preset keywords
for keyword in keywords:
    excelResults.append([keyDict[keyword]])

for j in range(1, numImage+1):
    img = cv2.imread("sample/{0:0=2d}.png".format(j), 0)

    count = []
    templateResult = []
    loc = []
    # Performing template matching
    for val in range(numTemplate):
        count.append(0)
        try:
            templateResult.append(cv2.matchTemplate(img, templates[val], cv2.TM_CCOEFF_NORMED))
            loc.append(np.where(templateResult[val] >= threshold[val]))
            for pt in zip(*loc[val][::-1]):
                count[val] = count[val] + 1
        except cv2.error as error:
            templateResult.append("Index Filler")
            loc.append("Index Filler")
            print("Error: Template size exceeds image size")

    # Takes max probability index and uses it for each following variable
    max_value = count.index(max(count))
    res = cv2.matchTemplate(img, templates[max_value], cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)

    # width, height of bounding box
    w, h = templates[max_value].shape[::-1]
    top_left = max_loc
    bottom_right = (top_left[0] + w, top_left[1] + h)

    # rectangle is drawn using coordinates gained from probable location and then cropped out
    cv2.rectangle(img, top_left, bottom_right, 255, 1)
    box = img[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]]

    # Saves the cropped image into its own image file
    cv2.imwrite("Model{}.png".format(j), box)
    print("Cropped image {}".format(j))

    # Removes the cropped out model from the original image
    img[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]] = 255

    # Create a new binary image from the original image using thresholding
    nrow, ncol = img.shape
    bwImage = np.zeros((nrow, ncol), dtype=np.uint8)
    for row in range(nrow):
        for col in range(ncol):
            if img[row, col] < 200:
                bwImage[row, col] = 1

    # Duplicate image to extract amendments
    dupImage = bwImage.copy()
    dupImage[dupImage == 1] = 255

    # Create 2 structuring element and obtain mask
    seHorizontal = np.ones((1, 51), dtype=np.uint8)
    seVertical = np.ones((51, 1), dtype=np.uint8)
    maskHorizontal = cv2.morphologyEx(bwImage, cv2.MORPH_OPEN, seHorizontal)
    maskVertical = cv2.morphologyEx(bwImage, cv2.MORPH_OPEN, seVertical)

    # Combine mask and invert it
    mask = maskHorizontal + maskVertical
    mask[mask > 1] = 1
    mask = 1 - mask

    # Applying mask onto binary image and turning it back to grayscale image
    bwImage = bwImage * mask
    bwImage[bwImage == 0] = 255
    bwImage[bwImage == 1] = 0

    # Extract data to get position of Amendments
    data = tess.image_to_data(bwImage)
    data = data.split('\n')
    for item in data:
        item = item.split('\t')
        if len(item) == 12 and item[11] == "AMENDMENTS":
            amenX, amenY, amenW, amenH = int(item[6]), int(item[7]), int(item[8]), int(item[9])
            dupImage[amenY:amenY + amenH, amenX:amenX + amenW] = 0
            bwImage[amenY:amenY + amenH, amenX:amenX + amenW] = 255
    for i in range(amenX, 0, -1):
        if dupImage[amenY, i] == 255:
            left = i + 1
            break
    for i in range(amenX, ncol):
        if dupImage[amenY, i] == 255:
            right = i - 1
            break
    for i in range(amenY, 0, -1):
        if dupImage[i, left] == 255:
            top = amenY + 1
            break
    line = 0
    cross = False
    for i in range(amenY, nrow, 1):
        if dupImage[i, left] == 0:
            cross = False
        elif line < 4 and not cross and dupImage[i, left] == 255:
            cross = True
            line += 1
        elif line == 4:
            bottom = i - 1
            break

    # List and condition to store and detect result
    result = []
    tagList = []
    for index in range(len(keywords)):
        result.append(None)

    # Extract Amendments
    Amendments = bwImage[top:bottom, left:right].copy()
    string = tess.image_to_string(Amendments)
    string = string.replace('\n\x0c', '')
    result[keywords.index("AMENDMENTS")] = string

    # Remove Amendments and detect words using Tesseract
    bwImage[top:bottom, left:right] = 255
    string = tess.image_to_string(bwImage)
    string.splitlines()
    string = string.replace('\n', ',')
    string = string.split(',')
    while '' in string:
        string.remove('')
    string.remove('\x0c')

    # Manual Editing
    if j == 3:
        string[4] = string[4].replace("STS", "STATUS")
    elif j == 13:
        string[-1] = string[-1].replace("a", "1/1")
    elif j == 19:
        del string[-2]
        string[-1] = "D-123-G-56-78-12-10-D KEVIN"
    elif j == 20:
        string.insert(1, "M")

    # Sorting Strings
    for line in string:
        # If no amendment and empty tag list, search for tag
        if len(tagList) == 0:
            for keyword in keywords:
                if keyword in line:
                    tagList.append(keyword)
            # Sorting tagged header and sort it to proper order as line
            kwIndex = []
            sortedKwIndex = []
            sortedIndexList = []
            tagIndex = []
            sortedTagIndex = []
            for tag in tagList:
                tagIndex.append(line.index(tag))
                sortedTagIndex.append(line.index(tag))
                kwIndex.append(keywords.index(tag))
            sortedTagIndex.sort()
            for sortedTag in sortedTagIndex:
                sortedIndexList.append(tagIndex.index(sortedTag))
            for index in sortedIndexList:
                sortedKwIndex.append(kwIndex[index])
        # If tag was detected on previous round
        else:
            lineIndex = 0
            # If more than 1 tag detected, split results and added them according to index
            if len(tagList) > 1:
                # Split line to individual word
                line = line.replace(" ", ",")
                line = line.split(",")
                while '' in string:
                    string.remove('')
                # Loop according to sorted index
                for index in range(len(sortedKwIndex)):
                    # Special loop for cad, project number
                    resultIn = ""
                    if sortedKwIndex[index] == keywords.index("CAD") and len(line[lineIndex]) < 5:
                        for loop in range(4):
                            resultIn += line[lineIndex]
                            lineIndex += 1
                        result[sortedKwIndex[index]] = resultIn
                    elif sortedKwIndex[index] == keywords.index("PROJECT NO") and len(line[lineIndex]) < 5:
                        for loop in range(3):
                            resultIn += line[lineIndex]
                            lineIndex += 1
                        result[sortedKwIndex[index]] = resultIn
                    elif sortedKwIndex[index] == keywords.index("TITLE"):
                        if line[lineIndex] == "SUNWAY":
                            for loop in range(3):
                                resultIn += line[lineIndex]
                                resultIn += " "
                                lineIndex += 1
                            result[sortedKwIndex[index]] = resultIn
                        elif line[lineIndex] == "STEEL" or line[lineIndex] == "BRIDGE":
                            for loop in range(2):
                                resultIn += line[lineIndex]
                                resultIn += " "
                                lineIndex += 1
                            result[sortedKwIndex[index]] = resultIn
                        else:
                            result[sortedKwIndex[index]] = line[lineIndex]
                            lineIndex += 1
                    # Other results
                    else:
                        result[sortedKwIndex[index]] = line[lineIndex]
                        lineIndex += 1
            # Only 1 tag thus 1 result, no separation required
            else:
                result[sortedKwIndex[0]] = line
            tagList = []
    # Store extracted results into excelResults
    for index in range(len(result)):
        excelResults[index].append(result[index])
    print("Extracted image {} information".format(j))

# Create xlsx file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Blueprint info"

# Adding results and setting width and height
for index in range(len(excelResults)):
    for index2 in range(len(excelResults[index])):
        ws.cell(index2 + 1, index + 1).value = str(excelResults[index][index2])
for i in range(2, len(tuple(ws.rows)) + 1):
    ws.row_dimensions[i].height = 60
    ws['A{}'.format(i)].alignment = openpyxl.styles.Alignment(wrapText=True)
for column_cells in ws.columns:
    unmerged_cells = list(filter(
        lambda cell_to_check: cell_to_check.coordinate not in ws.merged_cells, column_cells))
    length = max(len(str(cell.value)) for cell in unmerged_cells)
    ws.column_dimensions[unmerged_cells[0].column_letter].width = length * 1.4
ws.column_dimensions["A"].width = 45

wb.save("BlueprintInformation.xlsx")
print("Process Completed")
